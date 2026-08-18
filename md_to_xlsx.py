"""
Convert a Rota transcript (Markdown) into a styled .xlsx sheet.

Real transcript output varies a lot page to page — different column sets,
extra header rows inserted mid-grid, footers that don't all use the same
fields. Rather than fitting every page into one fixed template, this walks
the markdown top to bottom and renders each chunk in place:

  - a pipe-table block  -> a bordered, styled table
  - anything else       -> a plain text row (heading/bullet/paragraph)

Cell annotations from the transcription prompt's conventions are parsed out
of each table cell and turned into real Excel formatting:
  - [red] / [blue]                    -> font color
  - [strikethrough] / [crossed out] / [cross]  -> strikethrough font
  - [blank]                           -> empty cell
  - anything else in brackets (e.g. [covered], [correction: ...],
    [signature illegible]) -> kept as the visible value if the cell would
    otherwise be empty, and always added as a cell comment
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

TABLE_ROW_RE = re.compile(r"^\s*\|.*\|\s*$")
SEP_CELL_RE = re.compile(r"^:?-+:?$")
BRACKET_RE = re.compile(r"\[([^\[\]]+)]")

STRIKE_TAGS = {"strikethrough", "crossed out", "cross", "crossed"}
COLOR_TAGS = {"red", "blue"}


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def split_segments(md_text: str) -> list[tuple[str, list[str]]]:
    """Split markdown into ('table', lines) / ('text', lines) chunks, in order."""
    lines = md_text.splitlines()
    segments = []
    buffer: list[str] = []

    def flush():
        if buffer:
            segments.append(("text", buffer.copy()))
            buffer.clear()

    i = 0
    while i < len(lines):
        if TABLE_ROW_RE.match(lines[i]):
            flush()
            table_lines = []
            while i < len(lines) and TABLE_ROW_RE.match(lines[i]):
                table_lines.append(lines[i])
                i += 1
            segments.append(("table", table_lines))
        else:
            buffer.append(lines[i])
            i += 1
    flush()
    return segments


def _is_separator_row(cells: list[str]) -> bool:
    return bool(cells) and all(SEP_CELL_RE.fullmatch(c.strip()) for c in cells)


def parse_table_rows(table_lines: list[str]) -> list[list[str]]:
    """Turn raw '| a | b |' lines into a rectangular list of cell strings.

    Drops the markdown separator row (the '| :--- | :--- |' line) and pads
    every row to the widest row so the result stays rectangular even if the
    model produced a ragged table.
    """
    raw_rows = []
    for line in table_lines:
        inner = line.strip()
        if inner.startswith("|"):
            inner = inner[1:]
        if inner.endswith("|"):
            inner = inner[:-1]
        raw_rows.append([c.strip() for c in inner.split("|")])

    rows = [r for r in raw_rows if not _is_separator_row(r)]
    if not rows:
        return []
    max_cols = max(len(r) for r in rows)
    return [r + [""] * (max_cols - len(r)) for r in rows]


def parse_cell(raw: str) -> dict:
    """Extract display value + styling from one transcript table cell."""
    tags = [t.strip() for t in BRACKET_RE.findall(raw)]
    text = BRACKET_RE.sub("", raw).strip()

    bold = False
    if text.startswith("**") and text.endswith("**") and len(text) > 4:
        text = text[2:-2].strip()
        bold = True

    tags_lower = [t.lower() for t in tags]
    is_red = "red" in tags_lower
    is_blue = "blue" in tags_lower
    is_blank = "blank" in tags_lower
    is_strike = any(t in STRIKE_TAGS for t in tags_lower)
    comment_tags = [t for t in tags if t.lower() not in ("red", "blue", "blank")]

    if text:
        value = text
    elif is_blank:
        value = ""
    elif tags:
        value = f"[{tags[0]}]"  # e.g. [covered], [unclear], [signature illegible]
    else:
        value = ""

    return {
        "value": value,
        "bold": bold,
        "red": is_red,
        "blue": is_blue,
        "strike": is_strike,
        "comment": "; ".join(comment_tags) if comment_tags else None,
    }


def parse_text_line(line: str) -> dict:
    """Turn one non-table markdown line into display text + light styling."""
    text = line.strip()
    if not text:
        return {"value": "", "bold": False, "italic": False}

    heading = bool(re.match(r"^#+\s*", text))
    text = re.sub(r"^#+\s*", "", text)

    bullet = False
    if text.startswith("- "):
        text, bullet = text[2:].strip(), True
    elif text.startswith("* ") and not text.startswith("**"):
        text, bullet = text[2:].strip(), True

    bold = heading
    if text.startswith("**") and text.endswith("**") and len(text) > 4 and text.count("**") == 2:
        text, bold = text[2:-2].strip(), True
    else:
        # partial inline bold (e.g. "- **Center:** some text") — strip the
        # markers without forcing the whole line bold
        text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)

    italic = False
    if text.startswith("*") and text.endswith("*") and len(text) > 2 and not text.startswith("**"):
        text, italic = text[1:-1].strip(), True

    if text in ("---", ""):
        return {"value": "", "bold": False, "italic": False}

    return {"value": ("• " if bullet else "") + text, "bold": bold, "italic": italic}


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------

def _font(bold=False, italic=False, red=False, blue=False, strike=False):
    color = "FF0000" if red else ("0000FF" if blue else "000000")
    return Font(name=FONT_NAME, bold=bold, italic=italic, color=color, strike=strike)


def render_table(ws, start_row: int, rows: list[list[str]]) -> int:
    """Write one table starting at start_row. Returns the next free row."""
    if not rows:
        return start_row

    header, data_rows = rows[0], rows[1:]
    ncols = len(header)

    for j, raw in enumerate(header, start=1):
        parsed = parse_cell(raw)
        cell = ws.cell(row=start_row, column=j, value=parsed["value"] or raw.strip())
        cell.font = _font(bold=True, red=parsed["red"], blue=parsed["blue"])
        cell.alignment = CENTER
        cell.border = BORDER

    r = start_row + 1
    for row in data_rows:
        for j, raw in enumerate(row, start=1):
            parsed = parse_cell(raw)
            cell = ws.cell(row=r, column=j, value=parsed["value"])
            cell.font = _font(bold=parsed["bold"], red=parsed["red"], blue=parsed["blue"], strike=parsed["strike"])
            cell.alignment = CENTER
            cell.border = BORDER
            if parsed["comment"]:
                cell.comment = Comment(parsed["comment"], "Rota Transcript")
        r += 1
    return r + 1, ncols  # blank spacer row after the table


def render_text_block(ws, start_row: int, lines: list[str], width: int) -> int:
    r = start_row
    for line in lines:
        parsed = parse_text_line(line)
        if not parsed["value"]:
            r += 1
            continue
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(width, 1))
        cell = ws.cell(row=r, column=1, value=parsed["value"])
        cell.font = _font(bold=parsed["bold"], italic=parsed["italic"])
        cell.alignment = LEFT
        r += 1
    return r


def convert_transcript(md_path: Path, ws) -> None:
    """Render one transcript's markdown into the given worksheet, top to bottom."""
    md_text = Path(md_path).read_text(encoding="utf-8")
    segments = split_segments(md_text)

    # Pre-scan for the widest table so text rows merge across a sensible width.
    max_cols = 1
    for kind, lines in segments:
        if kind == "table":
            rows = parse_table_rows(lines)
            if rows:
                max_cols = max(max_cols, len(rows[0]))

    row = 1
    for kind, lines in segments:
        if kind == "table":
            rows = parse_table_rows(lines)
            if not rows:
                continue
            row, _ = render_table(ws, row, rows)
        else:
            row = render_text_block(ws, row, lines, max_cols)

    for col in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def sheet_name_for(md_path: Path) -> str:
    name = Path(md_path).stem
    return re.sub(r"[\[\]:*?/\\]", "_", name)[:31]


def convert_all(transcripts_dir: Path, output_path: Path) -> Path:
    """Convert every .md file in transcripts_dir into one workbook, one sheet per page."""
    transcripts_dir = Path(transcripts_dir)
    md_files = sorted(transcripts_dir.glob("*.md"))
    if not md_files:
        raise FileNotFoundError(f"No .md files found in {transcripts_dir}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default blank sheet
    for md_path in md_files:
        ws = wb.create_sheet(title=sheet_name_for(md_path))
        convert_transcript(md_path, ws)
        print(f"Converted {md_path.name} -> sheet '{ws.title}'")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import sys
    import config

    out = Path(sys.argv[1]) if len(sys.argv) > 1 else config.XLSX_DIR / "rota_transcripts.xlsx"
    convert_all(config.TRANSCRIPTS_DIR, out)
    print(f"Saved -> {out}")
